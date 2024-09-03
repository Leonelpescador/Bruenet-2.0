from .forms import ModificarPedidoForm  
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Caja, Pedido, Reserva, Inventario, Proveedor, Compra, Mesa, Pago, TransaccionCaja
from .forms import (
    AperturaCajaForm, 
    CierreCajaForm, 
    PedidoForm, 
    ModificarPedidoForm, 
    PagoForm, 
    ModificarPagoForm, 
    ReservaForm, 
    ModificarReservaForm, 
    CompraForm, 
    ModificarCompraForm, 
    MesaForm, 
    ProveedorForm, 
    InventarioForm
)
from django.contrib.auth.models import User

# Home View
@login_required
def home(request):
    caja_abierta = Caja.objects.filter(usuario=request.user, estado='abierta').first()
    
    mesas = Mesa.objects.all()

    context = {
        'mesas': mesas,
        'caja_abierta': caja_abierta,
    }

    return render(request, 'home.html', context)



# Creación de Pedido aca comienza
from .models import DetallePedido
from .forms import PedidoForm, DetallePedidoForm
from django.forms import inlineformset_factory

from django.forms import inlineformset_factory
from .models import Pedido, DetallePedido

# Formset para DetallePedido
DetallePedidoFormSet = inlineformset_factory(Pedido, DetallePedido, form=DetallePedidoForm, extra=1)


from django.http import JsonResponse
from .models import Categoria  # Importa el modelo de Categoría

@login_required
def crear_pedido(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    PedidoFormSet = inlineformset_factory(Pedido, DetallePedido, form=DetallePedidoForm, extra=1)
    categorias = Categoria.objects.all()  # Obtén todas las categorías

    if request.method == 'POST':
        pedido_form = PedidoForm(request.POST)
        formset = PedidoFormSet(request.POST)

        if pedido_form.is_valid() and formset.is_valid():
            pedido = pedido_form.save(commit=False)
            pedido.usuario = request.user
            pedido.mesa = mesa
            pedido.total = 0

            pedido.save()

            detalles = formset.save(commit=False)
            for detalle in detalles:
                detalle.pedido = pedido
                detalle.save()

            pedido.total = sum(detalle.subtotal for detalle in pedido.detalles.all())
            pedido.save()

            return JsonResponse({'success': True})

    else:
        pedido_form = PedidoForm()
        formset = PedidoFormSet()

    return render(request, 'pedido/crear_pedido.html', {
        'pedido_form': pedido_form,
        'formset': formset,
        'mesa': mesa,
        'categorias': categorias,  # Pasamos las categorías al contexto
    })


@login_required
def modificar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    PedidoFormSet = inlineformset_factory(Pedido, DetallePedido, form=DetallePedidoForm, extra=0)
    categorias = Categoria.objects.all()  # Mover la obtención de categorías fuera del bloque if

    if request.method == 'POST':
        pedido_form = ModificarPedidoForm(request.POST, instance=pedido)
        formset = PedidoFormSet(request.POST, instance=pedido)
        
        if pedido_form.is_valid() and formset.is_valid():
            pedido_form.save()
            formset.save()
            
            # Recalcular el total del pedido
            pedido.total = sum(item.subtotal for item in pedido.detalles.all())
            pedido.save()
            
            messages.success(request, 'Pedido modificado con éxito.')
            return redirect('home')
    else:
        pedido_form = ModificarPedidoForm(instance=pedido)
        formset = PedidoFormSet(instance=pedido)
    
    return render(request, 'pedido/modificar_pedido.html', {
        'pedido_form': pedido_form,
        'formset': formset,
        'pedido': pedido,
        'categorias': categorias,  # Pasamos las categorías al contexto
    })



@login_required
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        pedido.delete()
        messages.success(request, 'Pedido eliminado con éxito.')
        return redirect('pedidos_activos')
    
    return render(request, 'pedido/eliminar_pedido.html', {'pedido': pedido})

@login_required
def pedidos_activos(request):
    Pedido.objects.filter(en_proceso=True, usuario_procesando=request.user).update(en_proceso=False, usuario_procesando=None)
    pedidos = Pedido.objects.filter(estado__in=['pendiente', 'preparando', 'servido'])
    return render(request, 'pedido/pedidos_activos.html', {'pedidos': pedidos})


from django.http import JsonResponse
from .models import Menu

def obtener_precio_plato(request):
    menu_id = request.GET.get('menu_id')
    if menu_id:
        try:
            plato = Menu.objects.get(id=menu_id)
            return JsonResponse({'precio': plato.precio})
        except Menu.DoesNotExist:
            return JsonResponse({'error': 'Plato no encontrado'}, status=404)
    return JsonResponse({'error': 'Solicitud inválida'}, status=400)


#fin pedidos

# Creación de Pago
@login_required
def crear_pago(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.pedido = pedido
            pago.save()
            messages.success(request, 'Pago realizado con éxito.')
            return redirect('home')
    else:
        form = PagoForm()
    return render(request, 'pago/crear_pago.html', {'form': form, 'pedido': pedido})

# Modificación de Pago
@login_required
def modificar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        form = ModificarPagoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago modificado con éxito.')
            return redirect('home')
    else:
        form = ModificarPagoForm(instance=pago)
    return render(request, 'pago/modificar_pago.html', {'form': form})

# Eliminación de Pago
@login_required
def eliminar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        pago.delete()
        messages.success(request, 'Pago eliminado con éxito.')
        return redirect('home')
    return render(request, 'pago/eliminar_pago.html', {'pago': pago})




# Vista de Reservas
@login_required
def reservas(request):
    reservas = Reserva.objects.all()
    return render(request, 'reserva/reservas.html', {'reservas': reservas})

# Creación de Reserva


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .forms import ReservaForm, ModificarReservaForm
from .models import Reserva

# Creación de Reserva
@login_required
def crear_reserva(request):
    if request.method == 'POST':
        form = ReservaForm(request.POST)
        if form.is_valid():
            reserva = form.save(commit=False)  # No guardar aún en la base de datos
            reserva.usuario = request.user  # Asignar el usuario actual
            reserva.save()  # Ahora guardar en la base de datos
            messages.success(request, 'Reserva creada con éxito.')
            return redirect('reservas')  # Redirige a la lista de reservas o donde sea necesario
    else:
        form = ReservaForm()
    return render(request, 'reserva/crear_reserva.html', {'form': form})

#modifcar reserva
@login_required
def modificar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        form = ModificarReservaForm(request.POST, instance=reserva)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reserva modificada con éxito.')
            return redirect('reservas')
    else:
        form = ModificarReservaForm(instance=reserva)
    return render(request, 'reserva/editar_reserva.html', {'form': form})


# Eliminación de Reserva
@login_required
def eliminar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    if request.method == 'POST':
        reserva.delete()
        messages.success(request, 'Reserva eliminada con éxito.')
        return redirect('reservas')  # Redirige a la lista de reservas o donde sea necesario
    return render(request, 'reserva/eliminar_reserva.html', {'reserva': reserva})


#from de reservas  
from django import forms
from .models import Reserva

class ReservaForm(forms.ModelForm):
    fecha_reserva = forms.DateTimeField(
        widget=forms.DateTimeInput(attrs={'type': 'datetime-local'}),
        label='Fecha y Hora de la Reserva'
    )

    class Meta:
        model = Reserva
        fields = ['mesa', 'fecha_reserva', 'nombre_cliente', 'telefono_cliente', 'estado', 'numero_personas', 'comentarios']

class ModificarReservaForm(forms.ModelForm):
    fecha_reserva = forms.DateTimeField(
        widget=forms.DateTimeInput(attrs={'type': 'datetime-local'}),
        label='Fecha y Hora de la Reserva'
    )

    class Meta:
        model = Reserva
        fields = ['mesa', 'fecha_reserva', 'nombre_cliente', 'telefono_cliente', 'estado', 'numero_personas', 'comentarios']


#fin de flujo de reserva.






# Inventario
@login_required
def inventario(request):
    inventarios = Inventario.objects.all()
    return render(request, 'inventario/inventario.html', {'inventarios': inventarios})

# Creación de Inventario
def crear_inventario(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = InventarioForm()
    return render(request, 'inventario/crear_inventario.html', {'form': form})

# Edición de Inventario
def editar_inventario(request, pk):
    inventario = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = InventarioForm(instance=inventario)
    return render(request, 'inventario/editar_inventario.html', {'form': form})

# Eliminación de Inventario
def eliminar_inventario(request, pk):
    inventario = get_object_or_404(Inventario, pk=pk)
    if request.method == 'POST':
        inventario.delete()
        return redirect('inventario')
    return render(request, 'inventario/eliminar_inventario.html', {'inventario': inventario})


# Proveedores
@login_required
def proveedores(request):
    proveedores = Proveedor.objects.all()
    return render(request, 'proveedores/proveedores.html', {'proveedores': proveedores})

# Creación de Proveedor
@login_required
def crear_proveedor(request):
    if request.method == "POST":
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor creado exitosamente.')
            return redirect('proveedores')
        else:
            messages.error(request, 'Hubo un error al crear el proveedor. Por favor, verifica los datos ingresados.')
    else:
        form = ProveedorForm()

    return render(request, 'proveedores/crear_proveedores.html', {'form': form})  # Aquí debe coincidir con el nombre del archivo

# Edición de Proveedor
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from .forms import ProveedorForm

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == "POST":
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.has_changed():  # Verifica si se ha realizado algún cambio
            if form.is_valid():
                form.save()
                messages.success(request, 'Proveedor editado exitosamente.')
                return redirect('proveedores')
            else:
                messages.error(request, 'Hubo un error al editar el proveedor. Por favor, verifica los datos ingresados.')
        else:
            messages.info(request, 'No se han realizado cambios.')
            return redirect('proveedores')
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, 'proveedores/editar_proveedores.html', {'form': form})




# Eliminación de Proveedor
@login_required
def eliminar_proveedor(request, pk):
    proveedor = Proveedor.objects.get(pk=pk)
    if request.method == "POST":
        proveedor.delete()
        return redirect('proveedores')
    
    return render(request, 'proveedores/eliminar_proveedores.html', {'proveedor': proveedor})  # Aquí debe coincidir con el nombre del archivo


# Compras
from django.shortcuts import render, redirect
from .models import Compra
from .forms import CompraForm

def crear_compra(request):
    if request.method == 'POST':
        form = CompraForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('compras')
    else:
        form = CompraForm()
    
    return render(request, 'compra/crear_compra.html', {'form': form})

def compras(request):
    compras = Compra.objects.all()
    for compra in compras:
        # Determinamos si el archivo adjunto es un PDF
        if compra.archivo_documentacion:
            compra.es_pdf = compra.archivo_documentacion.url.lower().endswith('.pdf')
        else:
            compra.es_pdf = False
    
    return render(request, 'compra/compras.html', {'compras': compras})

@login_required
def editar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        form = CompraForm(request.POST, request.FILES, instance=compra)
        if form.is_valid():
            form.save()
            return redirect('compras')
    else:
        form = CompraForm(instance=compra)
    return render(request, 'compra/editar_compra.html', {'form': form})
@login_required
def eliminar_compra(request, pk):
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        compra.delete()
        return redirect('compras')
    return render(request, 'compra/eliminar_compra.html', {'compra': compra})
# Mesas
@login_required
def lista_mesas(request):
    mesas = Mesa.objects.all()
    return render(request, 'mesa/mesas.html', {'mesas': mesas})
@login_required
def crear_mesa(request):
    if request.method == 'POST':
        form = MesaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_mesas')
    else:
        form = MesaForm()
    return render(request, 'mesa/crear_mesa.html', {'form': form})
@login_required
def editar_mesa(request, pk):  
    mesa = get_object_or_404(Mesa, id=pk)
    if request.method == 'POST':
        form = MesaForm(request.POST, instance=mesa)
        if form.is_valid():
            form.save()
            return redirect('lista_mesas')
    else:
        form = MesaForm(instance=mesa)
    return render(request, 'mesa/editar_mesa.html', {'form': form})
@login_required
def eliminar_mesa(request, mesa_id):
    mesa = get_object_or_404(Mesa, id=mesa_id)
    mesa.delete()
    return redirect('lista_mesas')


# Flujo de Caja
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from .models import Caja, Pago, Pedido, TransaccionCaja
from django.contrib.auth.decorators import login_required

# Apertura de Caja
@login_required
def apertura_caja(request):
    caja_abierta = Caja.objects.filter(usuario=request.user, estado='abierta').first()
    if caja_abierta:
        messages.error(request, 'Ya tienes una caja abierta.')
        return redirect('consulta_caja', caja_id=caja_abierta.id)

    if request.method == 'POST':
        form = AperturaCajaForm(request.POST)
        if form.is_valid():
            caja = form.save(commit=False)
            caja.usuario = request.user  
            caja.save()
            messages.success(request, 'Caja abierta con éxito.')
            return redirect('consulta_caja', caja_id=caja.id)
    else:
        form = AperturaCajaForm()
    return render(request, 'caja/apertura_caja.html', {'form': form})



# Consulta de Caja
@login_required
def consulta_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id)
    

    transacciones = caja.transacciones.all()

    pedidos_en_caja = Pedido.objects.filter(
        fecha_pedido__gte=caja.apertura, 
        fecha_pedido__lte=caja.cierre if caja.cierre else timezone.now()
    )
    
    pagos = Pago.objects.filter(pedido__in=pedidos_en_caja)
    
    # Calcular el monto total de ingresos en la caja
    monto_total = transacciones.filter(tipo='ingreso').aggregate(Sum('monto'))['monto__sum'] or 0

    return render(request, 'caja/consulta_caja.html', {
        'caja': caja,
        'transacciones': transacciones,
        'monto_total': monto_total,
        'pagos': pagos
    })




# Cierre de Caja
@login_required
def cierre_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id, usuario=request.user)  # Asegura que solo pueda cerrar su propia caja

    if request.method == 'POST':
        total_final = caja.calcular_total_final()
        caja.cerrar_caja(total_final)
        messages.success(request, f'Caja {caja.id} cerrada exitosamente con un total de {caja.total_final}.')
        total_ingresos = sum(transaccion.monto for transaccion in caja.transacciones.filter(tipo='ingreso'))
        return render(request, 'caja/cierre_caja.html', {
            'caja': caja,
            'total_final': total_final,
            'total_ingresos': total_ingresos,
        })

    total_final = caja.calcular_total_final()
    total_ingresos = sum(transaccion.monto for transaccion in caja.transacciones.filter(tipo='ingreso'))
    
    return render(request, 'caja/cierre_caja.html', {
        'caja': caja,
        'total_final': total_final,
        'total_ingresos': total_ingresos,
    })





@login_required
def registrar_pago(request, pedido_id):
    caja_abierta = Caja.objects.filter(estado='abierta', usuario=request.user).first()
    if not caja_abierta:
        return redirect('apertura_caja')

    # Permitir registrar pago para pedidos en estado 'servido' o 'pendiente'
    pedido = get_object_or_404(Pedido, id=pedido_id, estado__in=['servido', 'pendiente'])

    # Verificar si el pedido está en proceso por otro usuario
    if pedido.en_proceso and pedido.usuario_procesando != request.user:
        messages.error(request, f"Este pedido está siendo procesado por {pedido.usuario_procesando.username}.")
        return redirect('pedidos_activos')
    
    # Marcar el pedido como en proceso y asignar el usuario actual
    pedido.marcar_en_proceso(request.user)

    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago')
        pago = Pago(pedido=pedido, metodo_pago=metodo_pago, monto=pedido.total)
        pago.save()
        pedido.estado = 'pagado'
        pedido.save()

        transaccion = TransaccionCaja(
            caja=caja_abierta,
            tipo='ingreso',
            monto=pago.monto,
            descripcion=f'Pago de pedido {pedido.id}'
        )
        transaccion.save()

        # Marcar el pedido como completado y liberar el proceso
        pedido.marcar_completado()

        return redirect('consulta_caja', caja_id=caja_abierta.id)
    
    return render(request, 'caja/registrar_pago.html', {'pedido': pedido, 'caja': caja_abierta})


@login_required
def marcar_servido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id, estado='preparando')

    if pedido.en_proceso and pedido.usuario_procesando != request.user:
        messages.error(request, f"Este pedido está siendo procesado por {pedido.usuario_procesando.username}.")
        return redirect('pedidos_activos')

    # Marcar el pedido como servido
    pedido.estado = 'servido'
    pedido.marcar_completado()  # Liberar el proceso si estaba marcado en proceso
    pedido.save()

    messages.success(request, f"El pedido {pedido.id} ha sido marcado como servido.")
    return redirect('pedidos_activos')

#fin





#reporte de cierre de caja 
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
import io
from django.utils import timezone

def descargar_reporte_caja(request, caja_id):
    caja = get_object_or_404(Caja, id=caja_id)
    template_path = 'caja/reporte_caja_pdf.html'
    context = {'caja': caja}
    
    # Verificar si la caja tiene una fecha de cierre
    if caja.cierre:
        fecha_cierre = caja.cierre.strftime("%Y-%m-%d_%H-%M-%S")
    else:
        # Si la caja no está cerrada, usar la fecha y hora actual
        fecha_cierre = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    nombre_archivo = f"Cierre_de_caja_{fecha_cierre}.pdf"
    
    # Renderizar la plantilla en un string
    html_string = render_to_string(template_path, context)
    
    # Crear un objeto de BytesIO para generar el PDF
    pdf = io.BytesIO()
    
    # Crear el PDF
    pisa_status = pisa.CreatePDF(
        io.BytesIO(html_string.encode('utf-8')),
        dest=pdf
    )
    
    # Si no hubo errores, preparar la respuesta con el PDF
    if not pisa_status.err:
        pdf.seek(0)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
    
    # Si hubo un error, mostrar el error en la respuesta
    return HttpResponse(f"Error al generar el PDF: {pisa_status.err}", content_type='text/plain')







# Logger

from django.contrib.auth import views as auth_views
class CustomLoginView(auth_views.LoginView):
    template_name = 'login.html'  


#pagina para clinetes.
from django.shortcuts import render
from .models import Categoria, Menu, Mesa

def cliente(request):
    # Obtener todas las mesas
    mesas = Mesa.objects.all()

    # Obtener todas las categorías y organizar los ítems de menú por categoría, filtrando solo los disponibles
    categorias = Categoria.objects.all()
    menu_items = {categoria.nombre: Menu.objects.filter(categoria=categoria, disponible=True) for categoria in categorias}

    return render(request, 'restaurant/cliente.html', {
        'mesas': mesas,
        'categorias': categorias,
        'menu_items': menu_items,
    })



#Menu "Platos."
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import MenuForm
from .models import Menu
from django.contrib.auth.decorators import login_required

@login_required
def listar_menu(request):
    # Agrupar los menús por categoría
    categorias = Menu.objects.values_list('categoria__nombre', flat=True).distinct()
    menu_categorias = {categoria: Menu.objects.filter(categoria__nombre=categoria, disponible=True) for categoria in categorias}
    menu_items_no_disponibles = Menu.objects.filter(disponible=False)

    return render(request, 'menu/listar_menu.html', {
        'menu_categorias': menu_categorias,
        'menu_items_no_disponibles': menu_items_no_disponibles
    })

from .models import Categoria
@login_required
def crear_menu(request):
    categorias = Categoria.objects.all()  # Obtener todas las categorías
    if request.method == 'POST':
        form = MenuForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'El menú ha sido guardado con éxito.')
            return redirect('listar_menu')
        else:
            messages.error(request, 'Error al guardar el menú. Verifica los datos ingresados.')
    else:
        form = MenuForm()

    return render(request, 'menu/crear_menu.html', {'form': form, 'categorias': categorias})

@login_required
def editar_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    categorias = Categoria.objects.all()  # Obtener todas las categorías
    
    if request.method == 'POST':
        form = MenuForm(request.POST, request.FILES, instance=menu_item)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Plato actualizado con éxito.')
            return redirect('listar_menu')
    else:
        form = MenuForm(instance=menu_item)
    
    return render(request, 'menu/editar_menu.html', {
        'form': form, 
        'menu': menu_item,
        'categorias': categorias  # Pasar las categorías al contexto
    })

@login_required
def eliminar_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    if request.method == 'POST':
        menu_item.delete()
        messages.success(request, 'Plato eliminado con éxito.')
        return redirect('listar_menu')
    return render(request, 'menu/eliminar_menu.html', {'menu': menu_item})

@login_required
def cambiar_disponibilidad_menu(request, menu_id):
    menu_item = get_object_or_404(Menu, id=menu_id)
    menu_item.disponible = not menu_item.disponible
    menu_item.save()
    messages.success(request, 'La disponibilidad del menú ha sido actualizada.')
    return redirect('listar_menu')

from django.http import HttpResponse
from .models import Menu

def filtrar_platos_por_categoria(request):
    categoria_id = request.GET.get('categoria_id')
    if categoria_id:
        platos = Menu.objects.filter(categoria_id=categoria_id, disponible=True)
        opciones = ''.join([f'<option value="{plato.id}">{plato.nombre_plato}</option>' for plato in platos])
        return HttpResponse(opciones)
    return HttpResponse('<option value="">No hay platos disponibles</option>')


#fin menu.


#pagos.

@login_required
def crear_pago(request):
    caja_abierta = Caja.objects.filter(estado='abierta').first()
    if not caja_abierta:
        messages.error(request, 'Debe abrir una caja antes de registrar pagos.')
        return redirect('apertura_caja')

    pedidos_pendientes = Pedido.objects.filter(estado='pendiente')

    if request.method == 'POST':
        form = PagoForm(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.caja = caja_abierta
            pago.save()
            messages.success(request, 'Pago registrado exitosamente.')
            return redirect('consulta_caja', caja_id=caja_abierta.id)
    else:
        form = PagoForm()

    return render(request, 'caja/crear_pago.html', {'form': form, 'pedidos': pedidos_pendientes})

@login_required
def modificar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        form = PagoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago modificado exitosamente.')
            return redirect('consulta_caja', caja_id=pago.caja.id)
    else:
        form = PagoForm(instance=pago)

    return render(request, 'caja/modificar_pago.html', {'form': form, 'pago': pago})

@login_required
def eliminar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    if request.method == 'POST':
        pago.delete()
        messages.success(request, 'Pago eliminado exitosamente.')
        return redirect('consulta_caja', caja_id=pago.caja.id)

    return render(request, 'caja/eliminar_pago.html', {'pago': pago})



from django.http import HttpResponseForbidden

@login_required
def mi_vista(request):
    if not request.user.has_perm('app_name.perm_name'):
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
    
    
#Procesos que tienen excel.     

import openpyxl
from django.http import HttpResponse

import openpyxl
from django.shortcuts import render
from django.contrib import messages



@login_required
def descargar_plantilla_proveedores(request):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Definir los encabezados de las columnas . aca es la estrucutra de la db. asi que ojo  si modifican le va a dar error chicos. 
    headers = ['Código', 'Nombre', 'CUIT', 'CBU', 'Alias CBU', 'Calle', 'N°', 'Localidad', 'País', 'Código Postal', 'Teléfono', 'Email', 'Plazo de Pago', 'Observaciones']
    ws.append(headers)

    # Configurar la respuesta HTTP para la descarga del archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_proveedores.xlsx'

    wb.save(response)
    return response


@login_required
def cargar_proveedores_masivo(request):
    if request.method == "POST":
        archivo_excel = request.FILES['archivo_excel']
        errores = []
        advertencias = []

        try:
            # Cargar el archivo Excel
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active

            # Verificar encabezados correctos
            headers = ['Código', 'Nombre', 'CUIT', 'CBU', 'Alias CBU', 'Calle', 'N°', 'Localidad', 'País', 'Código Postal', 'Teléfono', 'Email', 'Plazo de Pago', 'Observaciones']
            fila_encabezado = [cell.value for cell in ws[1]]
            
            if fila_encabezado != headers:
                mensajes = f"Encabezados incorrectos. Se esperaba: {headers} y se encontró {fila_encabezado}."
                messages.error(request, mensajes)
                return render(request, 'proveedores/cargar_proveedores.html')

            # Iterar sobre las filas del archivo Excel (saltando la primera fila que contiene los encabezados)
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                codigo, nombre, cuit, cbu, alias_cbu, calle, numero, localidad, pais, codigo_postal, telefono, email, plazo_pago, observaciones = row
                
                # Identificar columnas vacías y agregar advertencias
                columnas_vacias = []
                if not nombre:
                    columnas_vacias.append('Nombre')
                if not cuit:
                    columnas_vacias.append('CUIT')
                if not cbu:
                    columnas_vacias.append('CBU')
                if not alias_cbu:
                    columnas_vacias.append('Alias CBU')
                if not calle:
                    columnas_vacias.append('Calle')
                if not numero:
                    columnas_vacias.append('N°')
                if not localidad:
                    columnas_vacias.append('Localidad')
                if not pais:
                    columnas_vacias.append('País')
                if not codigo_postal:
                    columnas_vacias.append('Código Postal')
                if not telefono:
                    columnas_vacias.append('Teléfono')
                if not email:
                    columnas_vacias.append('Email')
                if not plazo_pago:
                    columnas_vacias.append('Plazo de Pago')
                if not observaciones:
                    columnas_vacias.append('Observaciones')

                if columnas_vacias:
                    advertencias.append(f"Fila {i}: Las siguientes columnas están vacías: {', '.join(columnas_vacias)}")

                # Si no hay errores, crear el proveedor
                if not errores:
                    Proveedor.objects.create(
                        codigo=codigo,
                        nombre=nombre,
                        cuit=cuit,
                        cbu=cbu,
                        alias_cbu=alias_cbu,
                        calle=calle,
                        numero=numero,
                        localidad=localidad,
                        pais=pais,
                        codigo_postal=codigo_postal,
                        telefono=telefono,
                        email=email,
                        plazo_pago=plazo_pago,
                        observaciones=observaciones
                    )

            if errores:
                messages.error(request, "Errores encontrados en el archivo:")
                return render(request, 'proveedores/cargar_proveedores.html', {'errores': errores})
            else:
                messages.success(request, "Proveedores cargados exitosamente.")
                if advertencias:
                    messages.warning(request, "Advertencias encontradas en el archivo:")
                    return render(request, 'proveedores/cargar_proveedores.html', {'advertencias': advertencias})

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
    
    return render(request, 'proveedores/cargar_proveedores.html')
